import pyperclip
import openpyxl
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from datetime import datetime
from enum import Enum
from fuzzywuzzy import process, fuzz
import heapq

class SearchMode(Enum):
    FUZZY = 0
    ACCURATE = 1

# 记忆碎片
class FragMem:
    def __init__(self, key, content, tagList=[]):
        self.key = key
        if content == '': # 如果输入框为空，则直接复制剪切板内容传入
            self.content = pyperclip.paste()
        else:
            self.content = content
        self.time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.tagList = ','.join(tagList)

class FragOps:
    def __init__(self, memFile, userName):
        self._userName = userName
        self.memFile = memFile
        # 打开记忆碎片存储文件
        try:
            self.wb = openpyxl.load_workbook(memFile)
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            self.wb.remove(self.wb.active)
            self.wb.create_sheet(title=f'{datetime.now().strftime("%Y")}')
            self.wb.save(memFile)
        # 创建线程池
        self.pool = ThreadPoolExecutor(max_workers=5)
    
    def newFrag(self, FragMem):
        year = FragMem.time[0:4]
        found = False
        i = 0
        for ws in self.wb.worksheets:
            if ws.title == year:
                found = True
                break
            else:
                i += 1
        # 没有当前年份的记忆表格，创建它
        if not found:
            self.wb.create_sheet(title=year)
        ws = self.wb.worksheets[i]
        ws.append([FragMem.key, FragMem.content, FragMem.tagList, self._userName, FragMem.time])
        self.wb.save(self.memFile)
    
    def delFrag(self, ws, index):
        ws.delete_rows(index)
        self.wb.save(self.memFile)

    def changeTagList(self, ws, index, change): # todo: after UI
        pass

    # 根据关键词模糊匹配 key 和 content，返回匹配度最高的记忆碎片
    # 搜索分为模糊搜索、精确搜索
    def search(self, keyword, mode=SearchMode.FUZZY, author=None, tagSet=set()):
        tList = [] # 管理搜索线程
        result = []
        lock = threading.Lock()
        searchFunList = {
            SearchMode.FUZZY: self.__fuzzySearchT,
            SearchMode.ACCURATE: self.__accurateSearchT
        }
        searchFun = searchFunList.get(mode)
        if searchFun is None:
            raise ValueError('Invalid mode')
        # 启动搜索线程，搜索所有表格
        for ws in self.wb.worksheets:
            t = self.pool.submit(searchFun, ws, keyword)
            tList.append(t)
        # 等待搜索结果返回
        for future in as_completed(tList):
            try:
                wsIndexList = future.result()
                with lock:
                    result.extend(wsIndexList)
            except Exception as e:
                print(f"An error occurred: {e}")
        # 筛选结果
        filterResult = self.__filter(result, author, tagSet)
        # test 打印搜索结果
        rowValuesList = []
        for i in range(0, len(filterResult), 2):
            worksheet = filterResult[i]
            indexList = filterResult[i+1]
            print(f'worksheet:{worksheet} indexList:{indexList}')
            for index in indexList:
                rowValues = [cell.value for cell in worksheet[index]]
                # print(rowValues)
                rowValuesList.append(rowValues)
        return rowValuesList

    # 基本搜索模式：模糊搜索
    # 按从大到小的顺序，返回worksheet里，和keyword匹配度最高的键值对的索引号
    def __fuzzySearchT(self, ws, keyword):
        ratioList = [] # 记录匹配度
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            ratioList.append(fuzz.partial_ratio(keyword, row))
        maxRatioList = heapq.nlargest(10, ratioList) # 获取n个最大匹配度的值
        indexList = [] # 记录最大匹配度内容对应的索引号
        for t in maxRatioList:
            if t < 20:
                continue
            index = ratioList.index(t)
            ratioList[index] = 0 # 不忽略匹配度相同的索引号
            index += 1
            indexList.append(index) # 存储index+1的值最为索引，而不是index，这是因为excel表索引是从1开始的
        return [ws, indexList]

    # 基本搜索模式：精确搜索
    # 返回完全匹配到的结果索引
    def __accurateSearchT(self, ws, keyword):
        indexList = []
        index = 1 # excel表索引是从1开始的
        for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
            if any(keyword in str(s).lower() for s in row):
                indexList.append(index)
            index += 1
        return [ws, indexList]

    # 过滤器
    def __filter(self, resultRaw, author=None, tagSet=set()):
        if not author and not tagSet: return resultRaw # 无筛选条件，直接返回
        result = []
        for i in range(0, len(resultRaw), 2): # 遍历所有列表
            ws = resultRaw[i] # worksheet
            indexList = resultRaw[i+1] # 未经筛选索引列表
            filterIndexList = [] # 筛选后的索引列表
            for index in indexList:
                rowValues = [cell.value for cell in ws[index]]
                # 作者不符合，跳过
                if author and author != rowValues[3]:
                    continue
                # tagList不符合，跳过
                if tagSet:
                    if rowValues[2]: # 空tagList，跳过
                        targetSet = set(rowValues[2].split(','))
                        if not tagSet.issubset(targetSet): # 不满足超集，跳过
                            continue
                    else:
                        continue
                # 符合筛选条件，保留
                filterIndexList.append(index)

            if filterIndexList:
                result.extend([ws, filterIndexList])
        return result